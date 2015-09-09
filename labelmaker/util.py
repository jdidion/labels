import datetime

class AliasedDictReader(object):
    """Wrapper around csv.DictReader that returns a dict in which keys are either
    column names (if header==True) or column indicies (of the form "col<idx>"). In
    the former case, rows are instances of AliasedDict, which enables items to be
    accessed either by column name or column index.
    """
    def __init__(self, infile, header=True, **kwargs):
        self.fh = open(infile, "rU")
        first_row = csv.reader(self.fh, delimiter=args.delimiter, skipinitialspace=True).next()
        aliases = list("col{0}".format(idx+1) for idx in xrange(len(first_row)))
        if header:
            fieldnames = first_row
            aliases = dict(zip(aliases, fieldnames))
        else:
            fieldnames = aliases
            aliases = None

        self.reader = csv.DictReader(i, fieldnames, **kwargs)
        self.aliases = aliases
    
    def __iter__(self):
        return self
    
    def next(self):
        row = self.reader.next()
        if self.aliases is not None:
            row = AliasedDict(row, self.aliases)
        return row
    
    def close(self):
        self.fh.close()

class ExcelReader(object):
    def __init__(self, infile, sheet, header=True, date_format="%Y-%m-%d"):
        """Open an excel workbook and iterate over the rows of a
        specific worksheet.
        i      -- a file name or file-like object
        sheet  -- sheet name or index
        header -- whether the workbook has a header row
        date_format -- strftime-style date format
        """
        from openpyxl import load_workbook
        
        wb = load_workbook(infile, data_only=True, use_iterators=True, keep_vba=False)
    
        try:
            sheet = wb.worksheets[int(sheet)-1]
    
        except:
            for ws in wb.worksheets:
                if ws.title == sheet:
                    sheet = ws
                    break
            else:
                raise Exception("No worksheet named {0}".format(sheet))
        
        self.iter = sheet.iter_rows()
        self.date_format = date_format
        
        if header:
            self.fieldnames = self.format_excel_row(self.iter.next())
            self.aliases = dict(zip(
                list("col{0}".format(idx+1) for idx in xrange(len(self.fieldnames))),
                self.fieldnames))
        else:
            self.fieldnames = list("col{0}".format(idx+1) for idx in xrange(len(sheet.columns))),
            self.aliases = None
    
    def __iter__(self):
        return self
    
    def next(self):
        row = self.iter.next()
        d = dict(zip(self.fieldnames, self.format_excel_row(row)))
        if self.aliases is not None:
            d = AliasedDict(d, self.aliases)
        return d
    
    def format_excel_row(self, row):
        def format_excel_cell(cell):
            val = cell.value
            if isinstance(val, datetime.datetime):
                return val.strftime(self.date_format)
            else:
                return str(val)
        return map(format_excel_cell, row)
    
    def close(self):
        pass

class AliasedDict(dict):
    def __init__(self, map_obj, aliases):
        dict.__init__(self, map_obj)
        self.aliases = aliases
    
    def __contains__(self, key):
        if super(AliasedDict, self).__contains__(key):
            return True
        else:
            return key in self.aliases
    
    def __missing__(self, key):
        if key in self.aliases:
            return self[self.aliases[key]]
        raise KeyError(key)

def safe_map(fn, seq):
    if isinstance(seq, tuple) or isinstance(seq, list):
        return map(fn, seq)
    else:
        return fn(seq)

def safe_get(seq, idx):
    if isinstance(seq, tuple) or isinstance(seq, list):
        return seq[idx]
    else:
        return seq
